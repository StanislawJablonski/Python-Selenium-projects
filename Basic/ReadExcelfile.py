import openpyxl

path = r"C:\Users\Young Stachu\Desktop\Python\miejsce na syf\excel.xlsx"

workbook = openpyxl.load_workbook(path)

sheet = workbook.active      #workbook.get_sheet_by_name("Sheet1")

rows = sheet.max_row
cols = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value)